import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
import openpyxl
import PySimpleGUI as sg

#Criação das telas de interação com usuário:
tela1 = [[sg.Text('Digite o nome ou caminho da planilha:')],
         [sg.Input(key='plan')],
         [sg.Text('Digite o número de Clusters:')],
         [sg.Input(key='key')],
         [sg.Button('Iniciar')]]

tela2 = [[sg.Text('O algoritmo K-means foi executado com sucesso.\nO resultado foi adicionado na planilha com sucesso.')],
         [sg.Button('Fechar')]]
tela3 = [[sg.Text('Planilha não encontrada')],
         [sg.Button('Fechar')]]

primeira_tela = sg.Window('K-means', layout= tela1)
segunda_tela = sg.Window('Obrigado', layout= tela2)
terceira_tela = sg.Window('Erro', layout= tela3)

while True:
    event,values = primeira_tela.read()
    if event == sg.WIN_CLOSED:
        break
    if event == 'Iniciar':
        nome = values['plan']
        key = values['key']
        primeira_tela.close()

#Define o caminho do arquivo Excel e a quantidade de grupos a serem analisados
k = int(key)
arquivo = nome + ".xlsx"

#Tratamento de erro caso o arquivo não seja encontrado
try:
    #Carrega o arquivo Excel em um DataFrame
    dataframe = pd.read_excel(arquivo)
    #Obtem os valores do DataFrame como uma matriz numpy
    dados = dataframe.values
    #Converte cada linha em um vetor distinto
    vetores = []
    
    for linha in dados:
        vetor = np.array(linha)
        vetores.append(vetor)

    #Imprime os vetores
    for vetor in vetores:
        print(vetor)
    
    #Executa o algoritmo K-means até a convergência            
    def kmeans(X, k):
        # Inicialização aleatória dos centróides
        centroids = X[np.random.choice(range(len(X)), size=k, replace=False)]
        
        while True:
            #Etapa de atribuição dos clusters
            labels = np.argmin(np.linalg.norm(X[:, np.newaxis] - centroids, axis=-1), axis=-1)
            
            #Atualização dos centróides
            new_centroids = np.array([X[labels == i].mean(axis=0) for i in range(k)])
            
            #Verificação de convergência
            if np.all(centroids == new_centroids):
                break
            
            centroids = new_centroids
        
        return centroids, labels

    #Atribue os dados da planilha (vetores) que serão analisados
    X = np.array(vetores)
    
    centroids, labels = kmeans(X, k)

    #Imprime os resultados no console
    print("Coordenadas dos centróides:")
    print(centroids)
    print("\nEtiquetas dos clusters:")
    print(labels)

    #Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(arquivo)

    #Seleciona a planilha onde deseja adicionar os dados
    nome_planilha = "Planilha1"
    planilha = workbook[nome_planilha]

    #Verifica a última coluna existente na planilha
    ultima_coluna = planilha.max_column

    #Obtem a letra da próxima coluna
    letra_coluna = openpyxl.utils.get_column_letter(ultima_coluna + 1)

    #Adicione os dados na próxima coluna disponível
    planilha[letra_coluna + str(1)] = "Grupo"
    for i, dado in enumerate(labels, start=2):
        planilha[letra_coluna + str(i)] = dado

    #Salva as alterações no arquivo Excel
    workbook.save(arquivo)
        
    #Cria um novo arquivo Excel para adicionar os valores das centroids
    novo_arquivo = nome + "_centroid.xlsx"
    workbook = openpyxl.Workbook()

    #Selecione a planilha
    planilha = workbook.active

    #Adicionar os dados das centróides na planilha
    for i, centroid in enumerate(centroids, start=1):
            for j, coordenada in enumerate(centroid, start=2):
                planilha.cell(row=i, column=j).value = coordenada

    #Salva o arquivo Excel
    workbook.save(novo_arquivo)
    print("Programa Finalizado")

    event = segunda_tela.read()
    if event == sg.WIN_CLOSED:
        segunda_tela.close()
    if event == 'Fechar':
        segunda_tela.close()
        
except:
    event = terceira_tela.read()
    if event == sg.WIN_CLOSED:
        terceira_tela.close()
    if event == 'Fechar':
        terceira_tela.close()
    print("Programa Finalizado")